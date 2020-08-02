import time

import openpyxl as openpyxl

# workSpace = openpyxl.Workbook()
# sheet = workSpace.create_sheet("观看信息", 0)
# sheet.cell(1, 1, "姓名")
# sheet.cell(1, 2, "手机号")
# sheet.cell(1, 3, "省份")
# sheet.cell(1, 4, "地市")
# sheet.cell(1, 5, "观看时长（分）")
# try:
#    i = 1
#    for i in range(1,1):
#        i += 1
#        sheet.cell(i, 1, "112")
# except:
#     print ("Error: unable to fetch data")
# workSpace.save("ad"+".xlsx")
t = int(time.time() * 1000)
print (t)                       #原始时间数据
t1  = (t + 3600*8 * 1000)/(3600*1000*24)
print(t1)