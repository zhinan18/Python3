import requests
import json
import pymysql
import openpyxl

# 打开数据库连接
db = pymysql.connect("localhost", "root", "werered", "livestream_yidong")
cursor = db.cursor()

eventId = input("eventId:")
i = 0
sql = "SELECT * FROM user_view_info \
       WHERE event_id = %s" % (eventId)
url = 'http://heshangwubeiyong.migucloud.com:8890/mobile/expData'
body = {"eventsId": eventId, "permisionType": "2"}
headers = {'content-type': "application/json"}
THREE_HOURS = 240

response = requests.post(url, data=json.dumps(body), headers=headers)
# 返回信息
resp = json.loads(response.text)
try:
    cursor.execute("DELETE from view_info")
    for userTime in resp['data']:
        if not userTime['timeCount']:
            userTime['timeCount'] = 0
        # 执行sql语句
        cursor.execute("INSERT INTO view_info(event_id, mobile, time_count)  VALUES ('%d', '%s',  %d)" % \
                       (int(eventId), userTime['mobile'], userTime['timeCount'] // 60))
        if userTime['mobile'].__eq__("15110175992"):
            print(userTime['mobile'], userTime['timeCount'])
        i += 1
    print(i)
    # 执行sql语句
    db.commit()
except:
    # 发生错误时回滚
    db.rollback()
    print("error db")

workSpace = openpyxl.Workbook()
sheet = workSpace.create_sheet("观看信息", 0)
sheet.cell(1, 1, "姓名")
sheet.cell(1, 2, "手机号")
sheet.cell(1, 3, "省份")
sheet.cell(1, 4, "地市")
sheet.cell(1, 5, "观看时长（分）")
try:
   # 执行SQL语句
   cursor.execute(sql)
   # 获取所有记录列表
   results = cursor.fetchall()
   i = 1
   for row in results:
       i += 1
       sheet.cell(i, 1, row[1])
       sheet.cell(i, 2, row[2])
       sheet.cell(i, 3, row[3])
       sheet.cell(i, 4, row[4])
       if row[5] > THREE_HOURS:
           sheet.cell(i, 5, THREE_HOURS)
       else:
           sheet.cell(i, 5, row[5])

except:
    print ("Error: unable to fetch data")
workSpace.save(eventId+".xlsx")

# 关闭数据库连接
db.close()
