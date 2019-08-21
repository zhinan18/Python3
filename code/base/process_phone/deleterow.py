import openpyxl
phoneList = openpyxl.load_workbook("201908161.xlsx")
sheet = phoneList.worksheets[0]
i = 0
j = 0
delList = list()
for cell in list(sheet.columns)[1]:
    j += 1
    phone = str(cell.value).strip(" ")
    if not phone.__eq__(cell.value) and not phone.__contains__("+"):
        print("black", phone, cell.value)
        sheet.cell(j, 2, phone)
    if len(phone) != 11:
        print(phone, j)
        delList.append(j-i)
        i += 1
print(i)
print(delList)
for row in delList:
    print(sheet.cell(row, 2).value)
    sheet.delete_rows(row)
phoneList.save("201908162.xlsx")

